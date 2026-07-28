# -*- coding: utf-8 -*-
"""Build 勤怠管理.xlsm : buttons (VBA) + auto-calculating records sheet."""
import os
import re
import shutil
import zipfile
import datetime as dt

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

import vbabuild

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, '勤怠管理.xlsm')
TMP_XLSX = os.path.join(HERE, '_tmp.xlsx')

REC = '勤務記録'          # records sheet display name
CLK = '打刻'             # timeclock sheet display name
NROWS = 300              # pre-filled formula rows (row 2 .. NROWS+1)

# ---------------------------------------------------------------------------
# VBA source
# ---------------------------------------------------------------------------
DOC_ATTR_WB = (
    'Attribute VB_Name = "ThisWorkbook"\r\n'
    'Attribute VB_Base = "0{00020819-0000-0000-C000-000000000046}"\r\n'
    'Attribute VB_GlobalNameSpace = False\r\n'
    'Attribute VB_Creatable = False\r\n'
    'Attribute VB_PredeclaredId = True\r\n'
    'Attribute VB_Exposed = True\r\n'
    'Attribute VB_TemplateDerived = False\r\n'
    'Attribute VB_Customizable = True\r\n'
)


def doc_attr_sheet(name):
    return (
        'Attribute VB_Name = "%s"\r\n'
        'Attribute VB_Base = "0{00020820-0000-0000-C000-000000000046}"\r\n'
        'Attribute VB_GlobalNameSpace = False\r\n'
        'Attribute VB_Creatable = False\r\n'
        'Attribute VB_PredeclaredId = True\r\n'
        'Attribute VB_Exposed = True\r\n'
        'Attribute VB_TemplateDerived = False\r\n'
        'Attribute VB_Customizable = True\r\n'
    ) % name


KINTAI_SRC = (
    'Attribute VB_Name = "Kintai"\r\n'
    'Option Explicit\r\n'
    '\r\n'
    'Private Const SHEET_NAME As String = "勤務記録"\r\n'
    '\r\n'
    "'―― 出勤ボタン ――\r\n"
    'Sub ClockIn()\r\n'
    '    Dim ws As Worksheet\r\n'
    '    Set ws = ThisWorkbook.Worksheets(SHEET_NAME)\r\n'
    '    Dim r As Long\r\n'
    '    r = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row\r\n'
    '    If r >= 2 Then\r\n'
    '        If ws.Cells(r, 2).Value <> "" And ws.Cells(r, 3).Value = "" Then\r\n'
    '            MsgBox "前回の退勤がまだ打刻されていません。" & vbCrLf & _\r\n'
    '                   "先に「退勤」を押してください。", vbExclamation, "出勤"\r\n'
    '            Exit Sub\r\n'
    '        End If\r\n'
    '    End If\r\n'
    '    If r < 1 Then r = 1\r\n'
    '    r = r + 1\r\n'
    '    ws.Cells(r, 1).Value = Date\r\n'
    '    ws.Cells(r, 2).Value = Time\r\n'
    '    ws.Cells(r, 1).NumberFormatLocal = "m/d"\r\n'
    '    ws.Cells(r, 2).NumberFormatLocal = "h:mm"\r\n'
    '    MsgBox Format(Date, "m/d") & "  " & Format(Time, "h:mm") & vbCrLf & _\r\n'
    '           "出勤を記録しました。", vbInformation, "出勤"\r\n'
    'End Sub\r\n'
    '\r\n'
    "'―― 退勤ボタン ――\r\n"
    'Sub ClockOut()\r\n'
    '    Dim ws As Worksheet\r\n'
    '    Set ws = ThisWorkbook.Worksheets(SHEET_NAME)\r\n'
    '    Dim r As Long\r\n'
    '    r = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row\r\n'
    '    If r < 2 Then\r\n'
    '        MsgBox "先に「出勤」を押してください。", vbExclamation, "退勤"\r\n'
    '        Exit Sub\r\n'
    '    End If\r\n'
    '    If ws.Cells(r, 2).Value = "" Then\r\n'
    '        MsgBox "出勤が打刻されていません。", vbExclamation, "退勤"\r\n'
    '        Exit Sub\r\n'
    '    End If\r\n'
    '    If ws.Cells(r, 3).Value <> "" Then\r\n'
    '        MsgBox "この行はすでに退勤済みです。" & vbCrLf & _\r\n'
    '               "新しく出勤する場合は「出勤」を押してください。", vbExclamation, "退勤"\r\n'
    '        Exit Sub\r\n'
    '    End If\r\n'
    '    ws.Cells(r, 3).Value = Time\r\n'
    '    ws.Cells(r, 3).NumberFormatLocal = "h:mm"\r\n'
    '    MsgBox Format(Time, "h:mm") & vbCrLf & "退勤を記録しました。", vbInformation, "退勤"\r\n'
    'End Sub\r\n'
)


def build_vba():
    modules = [
        vbabuild.Module('ThisWorkbook', DOC_ATTR_WB, is_document=True),
        vbabuild.Module('Sheet1', doc_attr_sheet('Sheet1'), is_document=True),
        vbabuild.Module('Sheet2', doc_attr_sheet('Sheet2'), is_document=True),
        vbabuild.Module('Kintai', KINTAI_SRC, is_document=False),
    ]
    return vbabuild.build('VBAProject', modules)


# ---------------------------------------------------------------------------
# Workbook (openpyxl)
# ---------------------------------------------------------------------------
def build_workbook():
    wb = openpyxl.Workbook()
    ws_clk = wb.active
    ws_clk.title = CLK
    ws_clk.sheet_properties.codeName = 'Sheet1'
    ws_rec = wb.create_sheet(REC)
    ws_rec.sheet_properties.codeName = 'Sheet2'

    thin = Side(style='thin', color='BBBBBB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    hdr_fill = PatternFill('solid', fgColor='D9D9D9')
    blue = Font(color='0000FF', bold=True)

    # ---- 打刻 sheet ----
    ws_clk['A1'] = '勤怠 打刻'
    ws_clk['A1'].font = Font(size=18, bold=True)
    ws_clk['A3'] = '下のボタンを押すだけで、「勤務記録」タブに時刻が記録され自動計算されます。'
    ws_clk['A4'] = '　① 仕事を始めるとき → 「出勤」ボタン'
    ws_clk['A5'] = '　② 仕事を終えたとき → 「退勤」ボタン'
    ws_clk['A6'] = '※ マクロを有効にしてください（ファイルを開いた際の警告バーで「コンテンツの有効化」）。'
    ws_clk['A6'].font = Font(color='C00000')
    ws_clk['A13'] = '※ ボタンが動かない環境（Web版Excel・Numbers・Googleスプレッドシート等）では、'
    ws_clk['A14'] = '　「勤務記録」タブに 日付・出勤・退勤 を直接入力しても自動計算されます。'
    for r in (13, 14):
        ws_clk['A%d' % r].font = Font(size=9, color='808080')
    ws_clk.column_dimensions['A'].width = 62

    # ---- 勤務記録 sheet ----
    headers = ['日付', '出勤', '退勤', '勤務時間', '給料']
    for c, h in enumerate(headers, start=1):
        cell = ws_rec.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = hdr_fill
        cell.border = border
        cell.alignment = center
    widths = {'A': 11, 'B': 9, 'C': 9, 'D': 14, 'E': 12, 'F': 10}
    for col, w in widths.items():
        ws_rec.column_dimensions[col].width = w

    # helper column F = 勤務分数 (integer minutes, rounded) -- hidden.
    # Rounding to whole minutes first avoids IEEE754 errors such as
    # 14:00->15:00 showing "0時間60分" instead of "1時間0分".
    ws_rec['F1'] = '分(数値)'
    ws_rec.column_dimensions['F'].hidden = True

    # formulas rows 2..NROWS+1
    for r in range(2, NROWS + 2):
        b, c = 'B%d' % r, 'C%d' % r
        f = 'F%d' % r
        ws_rec['F%d' % r] = ('=IF(OR(%s="",%s=""),"",ROUND(MOD(%s-%s,1)*1440,0))' % (b, c, c, b))
        ws_rec['D%d' % r] = (
            '=IF(OR(%s="",%s=""),"",INT(%s/60)&"時間 "&MOD(%s,60)&"分")'
            % (b, c, f, f))
        ws_rec['E%d' % r] = ('=IF(OR(%s="",%s=""),"",%s/480*$H$1)' % (b, c, f))
        ws_rec['A%d' % r].number_format = 'm/d'
        ws_rec['B%d' % r].number_format = 'h:mm'
        ws_rec['C%d' % r].number_format = 'h:mm'
        ws_rec['E%d' % r].number_format = '#,##0"円"'
        ws_rec['F%d' % r].number_format = '0'
        for col in range(1, 6):
            ws_rec.cell(row=r, column=col).border = border
            ws_rec.cell(row=r, column=col).alignment = center

    # sample rows from the reference image (year 2026)
    samples = [(dt.datetime(2026, 7, 24), dt.time(17, 30), dt.time(20, 0)),
               (dt.datetime(2026, 7, 25), dt.time(14, 0), dt.time(15, 0)),
               (dt.datetime(2026, 7, 26), dt.time(22, 0), dt.time(3, 0))]
    for i, (d, ti, to) in enumerate(samples):
        r = 2 + i
        ws_rec['A%d' % r] = d
        ws_rec['B%d' % r] = ti
        ws_rec['C%d' % r] = to

    # settings / totals block (right side)
    ws_rec['H1'] = 35000
    ws_rec['G1'] = '8時間あたりの給料(円)'
    ws_rec['H1'].font = blue
    ws_rec['H1'].number_format = '#,##0"円"'
    ws_rec['G2'] = '→ 時給(円)'
    ws_rec['H2'] = '=H1/8'
    ws_rec['H2'].number_format = '#,##0"円"'
    ws_rec['G4'] = '合計勤務時間'
    ws_rec['H4'] = '=SUM(F2:F%d)/60' % (NROWS + 1)
    ws_rec['H4'].number_format = '0.00"時間"'
    ws_rec['G5'] = '合計給料(円)'
    ws_rec['H5'] = '=SUM(E2:E%d)' % (NROWS + 1)
    ws_rec['H5'].number_format = '#,##0"円"'
    ws_rec['G7'] = '※ 上の3行はサンプルです。不要なら削除してください。'
    ws_rec['G7'].font = Font(size=9, color='808080')
    ws_rec['G8'] = '※ 黄色/青のセル(H1)を変えると給料の単価を調整できます。'
    ws_rec['G8'].font = Font(size=9, color='808080')
    ws_rec['H1'].fill = PatternFill('solid', fgColor='FFF2CC')
    ws_rec.column_dimensions['G'].width = 26
    ws_rec.column_dimensions['H'].width = 12

    ws_rec.freeze_panes = 'A2'
    wb.save(TMP_XLSX)


# ---------------------------------------------------------------------------
# Inject VBA + buttons into the xlsx zip -> xlsm
# ---------------------------------------------------------------------------
VML = '''<xml xmlns:v="urn:schemas-microsoft-com:vml" xmlns:o="urn:schemas-microsoft-com:office:office" xmlns:x="urn:schemas-microsoft-com:office:excel">
 <o:shapelayout v:ext="edit"><o:idmap v:ext="edit" data="1"/></o:shapelayout>
 <v:shapetype id="_x0000_t201" coordsize="21600,21600" o:spt="201" path="m,l,21600r21600,l21600,xe">
  <v:stroke joinstyle="miter"/>
  <v:path shadowok="f" o:extrusionok="f" gradientshapeok="t" o:connecttype="rect"/>
  <o:lock v:ext="edit" shapetype="t"/>
 </v:shapetype>
 <v:shape id="_x0000_s1025" type="#_x0000_t201" style='position:absolute;margin-left:15pt;margin-top:110pt;width:150pt;height:54pt;z-index:1;mso-wrap-style:tight' o:button="t" fillcolor="#4472c4" strokecolor="#2f528f">
  <v:fill color2="#4472c4" o:detectmouseclick="t"/>
  <o:lock v:ext="edit" rotation="t"/>
  <v:textbox style='mso-direction-alt:auto' o:singleclick="f">
   <div style='text-align:center'><font face="Meiryo" size="320" color="#FFFFFF"><b>&#20986;&#21220;</b></font></div>
  </v:textbox>
  <x:ClientData ObjectType="Button">
   <x:Anchor>0, 10, 6, 5, 2, 40, 9, 5</x:Anchor>
   <x:PrintObject>False</x:PrintObject>
   <x:AutoFill>False</x:AutoFill>
   <x:FmlaMacro>Kintai.ClockIn</x:FmlaMacro>
   <x:TextHAlign>Center</x:TextHAlign>
   <x:TextVAlign>Center</x:TextVAlign>
  </x:ClientData>
 </v:shape>
 <v:shape id="_x0000_s1026" type="#_x0000_t201" style='position:absolute;margin-left:180pt;margin-top:110pt;width:150pt;height:54pt;z-index:2;mso-wrap-style:tight' o:button="t" fillcolor="#c00000" strokecolor="#900000">
  <v:fill color2="#c00000" o:detectmouseclick="t"/>
  <o:lock v:ext="edit" rotation="t"/>
  <v:textbox style='mso-direction-alt:auto' o:singleclick="f">
   <div style='text-align:center'><font face="Meiryo" size="320" color="#FFFFFF"><b>&#36864;&#21220;</b></font></div>
  </v:textbox>
  <x:ClientData ObjectType="Button">
   <x:Anchor>2, 40, 6, 5, 5, 20, 9, 5</x:Anchor>
   <x:PrintObject>False</x:PrintObject>
   <x:AutoFill>False</x:AutoFill>
   <x:FmlaMacro>Kintai.ClockOut</x:FmlaMacro>
   <x:TextHAlign>Center</x:TextHAlign>
   <x:TextVAlign>Center</x:TextVAlign>
  </x:ClientData>
 </v:shape>
</xml>'''


def inject():
    vba = build_vba()
    zin = zipfile.ZipFile(TMP_XLSX, 'r')
    names = zin.namelist()
    data = {n: zin.read(n) for n in names}
    zin.close()

    # --- [Content_Types].xml ---
    ct = data['[Content_Types].xml'].decode('utf-8')
    if 'Extension="bin"' not in ct:
        ct = ct.replace('</Types>',
            '<Default Extension="bin" ContentType="application/vnd.ms-office.vbaProject"/></Types>')
    if 'Extension="vml"' not in ct:
        ct = ct.replace('</Types>',
            '<Default Extension="vml" ContentType="application/vnd.openxmlformats-officedocument.vmlDrawing"/></Types>')
    ct = ct.replace(
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml',
        'application/vnd.ms-excel.sheet.macroEnabled.main+xml')
    data['[Content_Types].xml'] = ct.encode('utf-8')

    # --- workbook.xml : codeName on workbookPr ---
    wbxml = data['xl/workbook.xml'].decode('utf-8')
    if '<workbookPr' in wbxml:
        if 'codeName=' not in wbxml.split('<workbookPr', 1)[1].split('>', 1)[0]:
            wbxml = wbxml.replace('<workbookPr', '<workbookPr codeName="ThisWorkbook"', 1)
    else:
        wbxml = re.sub(r'(<workbook[^>]*>)', r'\1<workbookPr codeName="ThisWorkbook"/>', wbxml, count=1)
    data['xl/workbook.xml'] = wbxml.encode('utf-8')

    # --- workbook rels : vbaProject ---
    wr = data['xl/_rels/workbook.xml.rels'].decode('utf-8')
    if 'vbaProject.bin' not in wr:
        wr = wr.replace('</Relationships>',
            '<Relationship Id="rIdVba" '
            'Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" '
            'Target="vbaProject.bin"/></Relationships>')
    data['xl/_rels/workbook.xml.rels'] = wr.encode('utf-8')

    # --- sheet1 (打刻) : legacyDrawing ---
    s1 = data['xl/worksheets/sheet1.xml'].decode('utf-8')
    if 'xmlns:r=' not in s1[:200]:
        s1 = s1.replace('<worksheet ',
            '<worksheet xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" ', 1)
    s1 = s1.replace('</worksheet>', '<legacyDrawing r:id="rIdVml"/></worksheet>')
    data['xl/worksheets/sheet1.xml'] = s1.encode('utf-8')

    # sheet1 rels
    s1rels_name = 'xl/worksheets/_rels/sheet1.xml.rels'
    if s1rels_name in data:
        s1r = data[s1rels_name].decode('utf-8')
        s1r = s1r.replace('</Relationships>',
            '<Relationship Id="rIdVml" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing" '
            'Target="../drawings/vmlDrawing1.vml"/></Relationships>')
    else:
        s1r = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
               '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
               '<Relationship Id="rIdVml" '
               'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing" '
               'Target="../drawings/vmlDrawing1.vml"/></Relationships>')
    data[s1rels_name] = s1r.encode('utf-8')

    # --- new parts ---
    data['xl/drawings/vmlDrawing1.vml'] = VML.encode('utf-8')
    data['xl/vbaProject.bin'] = vba

    # --- write xlsm ---
    if os.path.exists(OUT):
        os.remove(OUT)
    zout = zipfile.ZipFile(OUT, 'w', zipfile.ZIP_DEFLATED)
    for n, d in data.items():
        zout.writestr(n, d)
    zout.close()


if __name__ == '__main__':
    build_workbook()
    inject()
    os.remove(TMP_XLSX)
    print('wrote', OUT, os.path.getsize(OUT), 'bytes')
